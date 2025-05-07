from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from .models import DersProgramiSlotu, Bolum, Ders, OgretimUyesi, Derslik
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import datetime
import pandas as pd # Pandas'ı import edelim
from django.db import IntegrityError
from .scheduler import BacktrackingScheduler # Scheduler'ı import et
from django.core.management import call_command
import logging
import json
from django.views.decorators.http import require_POST

# Logger'ı modül seviyesinde tanımla
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s') # Temel ayarlar (isteğe bağlı)

# Create your views here.

# Zaman aralıklarını ve günleri tanımlayalım
TIME_SLOTS = [
    (datetime.time(8, 0), datetime.time(9, 0)),
    (datetime.time(9, 0), datetime.time(10, 0)),
    (datetime.time(10, 0), datetime.time(11, 0)),
    (datetime.time(11, 0), datetime.time(12, 0)),
    (datetime.time(12, 0), datetime.time(13, 0)),
    (datetime.time(13, 0), datetime.time(14, 0)),
    (datetime.time(14, 0), datetime.time(15, 0)),
    (datetime.time(15, 0), datetime.time(16, 0)),
    (datetime.time(16, 0), datetime.time(17, 0)),
    # İsterlerde online dersler için ek saatler vardı, onları da ekleyelim
    (datetime.time(17, 0), datetime.time(19, 0)), 
    (datetime.time(19, 0), datetime.time(21, 0)),
]
DAYS = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
SINIFLAR = [1, 2, 3, 4]

# Zaman aralıkları ve günler (Excel export ve scheduler ile aynı olmalı)
# Ortak bir yerde tanımlamak daha iyi olabilir (örn: schedule/constants.py)
TIME_SLOTS_DISPLAY = [
    (datetime.time(8, 0), datetime.time(9, 0)),
    (datetime.time(9, 0), datetime.time(10, 0)),
    (datetime.time(10, 0), datetime.time(11, 0)),
    (datetime.time(11, 0), datetime.time(12, 0)),
    (datetime.time(12, 0), datetime.time(13, 0)),
    (datetime.time(13, 0), datetime.time(14, 0)),
    (datetime.time(14, 0), datetime.time(15, 0)),
    (datetime.time(15, 0), datetime.time(16, 0)),
    (datetime.time(16, 0), datetime.time(17, 0)),
    (datetime.time(17, 0), datetime.time(19, 0)), 
    (datetime.time(19, 0), datetime.time(21, 0)),
]
DAYS_DISPLAY = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
SINIFLAR_DISPLAY = [1, 2, 3, 4]

def export_schedule_excel(request):
    # --- Filtreleri Al (view_schedule'dan benzer kod) ---
    selected_bolum_id = request.GET.get('bolum')
    selected_sinif = request.GET.get('sinif')
    selected_semester = request.GET.get('semester') # Dönem filtresini al

    # Tüm bölümleri ve sınıfları tanımla (varsayılanlar)
    all_bolumler = Bolum.objects.all()
    all_siniflar = SINIFLAR_DISPLAY
    all_semesters = ["Güz", "Bahar"] # Dönem seçenekleri

    # Temel sorguyu oluştur
    slots = DersProgramiSlotu.objects.select_related('ders', 'ogretim_uyesi', 'derslik', 'bolum').order_by('gun', 'baslangic_saati', 'bolum', 'sinif')

    # Filtreleri uygula
    current_bolum = None
    bolum_adi_baslik = "Tüm Bölümler"
    siniflar_to_display = list(all_siniflar) # Başlangıçta tüm sınıfları göster

    if selected_bolum_id:
        try:
            slots = slots.filter(bolum_id=selected_bolum_id)
            current_bolum = all_bolumler.get(id=selected_bolum_id)
            bolum_adi_baslik = current_bolum.bolum_adi
        except (Bolum.DoesNotExist, ValueError):
             selected_bolum_id = None # Geçersiz ID ise filtreyi temizle

    if selected_sinif:
        try:
            sinif_int = int(selected_sinif)
            if sinif_int in all_siniflar:
                slots = slots.filter(sinif=sinif_int)
                siniflar_to_display = [sinif_int] # Sadece seçili sınıfı göster
                if current_bolum:
                     bolum_adi_baslik += f" - {sinif_int}. Sınıf"
                else:
                     bolum_adi_baslik = f"{sinif_int}. Sınıf (Tüm Bölümler)"
            # else: Geçersiz sınıfsa tüm sınıfları göstermeye devam et
        except ValueError: # Sınıf sayı değilse tümünü göster
            pass
    elif current_bolum: # Sadece bölüm seçilmişse başlığa ekle
         bolum_adi_baslik += " (Tüm Sınıflar)"
         
    # Dönem filtresini uygula
    if selected_semester and selected_semester in all_semesters:
        slots = slots.filter(semester=selected_semester)
        # Başlığa dönem bilgisini ekle (Excel başlığı için)
        if bolum_adi_baslik == "Tüm Bölümler":
            bolum_adi_baslik = f"{selected_semester} Dönemi (Tüm Bölümler)"
        elif selected_sinif: # Eğer sınıf da seçiliyse
             bolum_adi_baslik += f" - {selected_semester}"
        elif current_bolum: # Sadece bölüm seçiliyse
             bolum_adi_baslik = f"{current_bolum.bolum_adi} - {selected_semester} Dönemi"
    # else: # Boş veya geçersizse filtreleme yok, başlık değişmez

    # --- Veri Hazırlama (Filtrelenmiş slotlarla) --- 
    schedule_data = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
    
    for slot in slots:
        # Slotun başlangıç saatine uygun zaman dilimini bul
        slot_time_str = f"{slot.baslangic_saati.strftime('%H:%M')}-{slot.bitis_saati.strftime('%H:%M')}"
        # Ders bilgilerini formatla (Dönem bilgisi eklendi)
        ders_info = f"{slot.ders.ders_kodu}\n{slot.ogretim_uyesi.ad_soyad}\n{slot.derslik.derslik_adi}\n({slot.semester})" 
        # Sadece gösterilecek sınıflar için veri ekle
        if slot.sinif in siniflar_to_display:
             schedule_data[slot.gun][slot_time_str][slot.sinif] = ders_info

    # --- Excel Oluşturma --- 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Haftalık Ders Programı"

    # --- Stil Tanımları ---
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    header_font = Font(bold=True, name='Calibri', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    # Sınıf renkleri (örnek)
    renkler = {
        1: PatternFill(start_color="FFFFE0B2", end_color="FFFFE0B2", fill_type="solid"), # Açık Turuncu (1. Sınıf)
        2: PatternFill(start_color="FFBBDEFB", end_color="FFBBDEFB", fill_type="solid"), # Açık Mavi (2. Sınıf)
        3: PatternFill(start_color="FFFFF9C4", end_color="FFFFF9C4", fill_type="solid"), # Açık Sarı (3. Sınıf)
        4: PatternFill(start_color="FFC8E6C9", end_color="FFC8E6C9", fill_type="solid"), # Açık Yeşil (4. Sınıf)
    }
    gun_fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid") # Gri

    # --- Başlıkları Yazma (Yeni Düzen) --- 
    # Satır 1: Bölüm Başlığı (Filtreye göre)
    ws.cell(row=1, column=1).border = thin_border # A1
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + len(DAYS_DISPLAY))
    ws.cell(row=1, column=2, value=bolum_adi_baslik).font = header_font 
    ws.cell(row=1, column=2).alignment = center_alignment
    for col_idx in range(2, 2 + len(DAYS_DISPLAY)):
        ws.cell(row=1, column=col_idx).border = thin_border

    # Satır 2: Saatler başlığı ve Gün Başlıkları
    ws.cell(row=2, column=1, value="Saatler").font = header_font
    ws.cell(row=2, column=1).alignment = center_alignment
    ws.cell(row=2, column=1).border = thin_border

    for j, day in enumerate(DAYS_DISPLAY):
        col = 2 + j # Sütun 2'den başla (B)
        ws.cell(row=2, column=col, value=day).font = header_font
        ws.cell(row=2, column=col).alignment = center_alignment
        ws.cell(row=2, column=col).border = thin_border
        ws.cell(row=2, column=col).fill = gun_fill # Gün başlıklarına gri dolgu

    # --- Verileri Yazma (Yeni Düzen) --- 
    time_slots_str_list = [f"{ts[0].strftime('%H:%M')}-{ts[1].strftime('%H:%M')}" for ts in TIME_SLOTS_DISPLAY]
    
    for i, time_str in enumerate(time_slots_str_list):
        current_row = 3 + i # Satır 3'ten başla
        # İlk sütuna saat aralığını yaz
        ws.cell(row=current_row, column=1, value=time_str).alignment = center_alignment
        ws.cell(row=current_row, column=1).border = thin_border

        # Günler için sütunlara dersleri yaz
        for j, day in enumerate(DAYS_DISPLAY):
            col = 2 + j # Sütun 2'den başla (B)
            # Bu gün ve saatteki dersleri al (birden fazla sınıf olabilir)
            cell_content = []
            for sinif in siniflar_to_display:
                ders_info = schedule_data[day][time_str].get(sinif, "")
                if ders_info:
                    # Eğer birden fazla sınıf gösteriliyorsa, sınıfı başına ekle
                    if len(siniflar_to_display) > 1:
                        cell_content.append(f"{sinif}. Sınıf:\n{ders_info}")
                    else:
                        cell_content.append(ders_info)
            
            # Hücreye birleştirilmiş içeriği yaz
            final_cell_value = "\n---\n".join(cell_content) # Farklı sınıfları ayırıcıyla birleştir
            cell = ws.cell(row=current_row, column=col, value=final_cell_value)
            cell.alignment = center_alignment # İçeriği ortala ve sar
            cell.border = thin_border
            # Sınıf rengi yerine, belki buraya renk eklememek daha iyi olur
            # Veya sadece ilk dersin sınıfına göre renklendirilebilir
            # Şimdilik renksiz bırakalım

    # --- Sütun Genişliklerini Ayarlama (Yeni Düzen) ---
    ws.column_dimensions[get_column_letter(1)].width = 15 # Saat sütunu (A)
    for j in range(len(DAYS_DISPLAY)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 30 # Gün sütunları daha geniş

    # --- Yanıtı Hazırlama --- 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=ders_programi.xlsx'
    wb.save(response)

    return response

@staff_member_required # Sadece admin/staff kullanıcılar erişebilsin
def import_courses_view(request):
    if request.method == 'POST':
        try:
            bolum_id = request.POST.get('bolum')
            file = request.FILES.get('file')
            default_kontenjan = int(request.POST.get('default_kontenjan', 50))

            if not bolum_id or not file:
                messages.error(request, "Bölüm seçimi ve dosya yükleme zorunludur.")
                return redirect(request.path) # Aynı sayfaya geri dön

            bolum = Bolum.objects.get(id=bolum_id)

            # Dosya uzantısına göre pandas ile oku
            try:
                if file.name.endswith('.csv'):
                    df = pd.read_csv(file, encoding='utf-8-sig')
                elif file.name.endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(file)
                else:
                    messages.error(request, "Desteklenmeyen dosya formatı. Lütfen CSV veya XLSX kullanın.")
                    return redirect(request.path)
            except Exception as e:
                 messages.error(request, f"Dosya okunurken hata oluştu: {e}")
                 return redirect(request.path)
            
            # Sütun isimlerini kontrol et (büyük/küçük harf duyarsız yapabiliriz)
            # Kredi ve Akts isteğe bağlı olabilir, Ders Tipi önemli.
            required_columns = ['Dönem', 'Ders Kodu', 'Ders Adi', 'Teorik', 'Uygulama', 'Laboratuar', 'Ders Tipi']
            optional_columns = ['Kredi', 'Akts']
            df.columns = df.columns.str.strip() # Başındaki/sonundaki boşlukları temizle
            
            # Sütun adlarını normalleştirelim (küçük harf, boşluk yerine alt çizgi)
            df.columns = df.columns.str.lower().str.replace(' ', '_').str.replace('ı','i')
            # Beklenen sütun adlarını da normalleştirelim
            expected_required = [col.lower().replace(' ', '_').replace('ı','i') for col in required_columns]
            expected_optional = [col.lower().replace(' ', '_').replace('ı','i') for col in optional_columns]

            if not all(col in df.columns for col in expected_required):
                missing = [req for req in required_columns if req.lower().replace(' ', '_').replace('ı','i') not in df.columns]
                messages.error(request, f"Dosya gerekli sütunları içermiyor: {missing}")
                return redirect(request.path)

            created_count = 0
            updated_count = 0
            skipped_count = 0

            for index, row in df.iterrows():
                try:
                    # Veriyi işle (normalleştirilmiş sütun adlarını kullan)
                    dönem_str = str(row.get('dönem', '')).split('.')[0]
                    ders_kodu = str(row.get('ders_kodu', '')).strip()
                    ders_adi = str(row.get('ders_adi', '')).strip()
                    ders_tipi_str = str(row.get('ders_tipi', 'Zorunlu')).strip().lower() # Zorunlu/Seçmeli bilgisi
                    teorik_str = str(row.get('teorik', '0')).split('.')[0]
                    uygulama_str = str(row.get('uygulama', '0')).split('.')[0]
                    lab_str = str(row.get('laboratuar', '0')).split('.')[0]
                    kredi_str = str(row.get('kredi', '0')).replace(',', '.') # Ondalık ayıracı nokta yap
                    akts_str = str(row.get('akts', '0')).split('.')[0]

                    if not dönem_str or not dönem_str.isdigit() or not ders_kodu or not ders_adi:
                        skipped_count += 1
                        continue
                        
                    # Dönemi al
                    donem = int(dönem_str)
                    # Sınıfı hesapla
                    sinif = (donem + 1) // 2
                    
                    teorik = int(teorik_str) if teorik_str.isdigit() else 0
                    uygulama = int(uygulama_str) if uygulama_str.isdigit() else 0
                    lab = int(lab_str) if lab_str.isdigit() else 0
                    
                    # Kredi ve AKTS'yi güvenli bir şekilde dönüştür
                    try:
                        kredi = float(kredi_str) if kredi_str else None
                    except ValueError:
                        kredi = None
                    try:
                        akts = int(akts_str) if akts_str.isdigit() else None
                    except ValueError:
                         akts = None

                    haftalik_saat = teorik + uygulama + lab
                    # Haftalık saati 0 olan zorunlu dersleri atla (Seçmeli dersler 0 saat olabilir)
                    if haftalik_saat == 0 and ders_tipi_str == 'zorunlu': 
                        skipped_count += 1
                        continue
                            
                    # Yapısal ders tipini belirle (Lab > Uygulama > Teorik önceliği)
                    if lab > 0:
                        yapısal_ders_tipi = 'LAB'
                    elif uygulama > 0:
                        yapısal_ders_tipi = 'UYGULAMA'
                    else:
                        yapısal_ders_tipi = 'TEORIK'
                        
                    # Zorunlu mu?
                    is_zorunlu = ders_tipi_str == 'zorunlu'

                    # Ortak ders mi? (TUR, ATA, DIL kodları ile başlayanlar)
                    shared_prefixes = ['TUR', 'ATA', 'DIL']
                    is_shared_course = any(ders_kodu.upper().startswith(prefix) for prefix in shared_prefixes)

                    # Dersin unique anahtarını belirle (Ortaksa sadece kod, değilse kod+bölüm)
                    if is_shared_course:
                        lookup_keys = {'ders_kodu': ders_kodu}
                    else:
                        lookup_keys = {'ders_kodu': ders_kodu, 'bolum': bolum}

                    defaults = {
                        'ders_adi': ders_adi,
                        'bolum': bolum, # defaults içinde bölüm her zaman güncellenir
                        'sinif': sinif,
                        'donem': donem,
                        'haftalik_saat': haftalik_saat,
                        'tip': yapısal_ders_tipi,
                        'kontenjan': default_kontenjan,
                        'is_zorunlu': is_zorunlu,
                        'kredi': kredi,
                        'akts': akts,
                        'is_shared': is_shared_course
                    }

                    # update_or_create'i dinamik anahtarlarla çağır
                    obj, created = Ders.objects.update_or_create(
                        **lookup_keys, # Anahtarları burada aç
                        defaults=defaults
                    )
                        
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except (ValueError, TypeError, IntegrityError) as e:
                     messages.warning(request, f'Satır {index + 2} işlenirken hata: {row.to_dict()} - Hata: {e}')
                     skipped_count += 1
                except Exception as e:
                     messages.error(request, f'Satır {index + 2} işlenirken beklenmedik hata: {row.to_dict()} - Hata: {e}')
                     skipped_count += 1
            
            messages.success(request, f'{created_count} ders başarıyla eklendi, {updated_count} ders güncellendi, {skipped_count} ders atlandı.')

            # --- OTOMATİK HOCA ATAMA --- #
            try:
                logger.info("Import successful. Triggering automatic instructor assignment...")
                call_command('assign_all_instructors') # Yönetim komutunu çağır
                messages.info(request, "Öğretim üyesi atamaları otomatik olarak çalıştırıldı.")
            except Exception as assign_error:
                logger.error(f"Error during automatic instructor assignment: {assign_error}")
                messages.warning(request, f"Dersler içe aktarıldı ancak otomatik öğretim üyesi atama sırasında bir hata oluştu: {assign_error}")
            # --- OTOMATİK HOCA ATAMA SONU --- #

            # Başarılı içe aktarma sonrası Ders listeleme sayfasına yönlendir
            return redirect('admin:schedule_ders_changelist')

        except Bolum.DoesNotExist:
             messages.error(request, "Seçilen bölüm bulunamadı.")
             return redirect(request.path)
        except Exception as e:
            messages.error(request, f"Beklenmedik bir hata oluştu: {e}")
            return redirect(request.path)

    else: # GET request
        bolumler = Bolum.objects.all()
        context = {
            'title': 'Dersleri İçe Aktar',
            'bolumler': bolumler,
            'opts': Ders._meta, # Admin template'inin çalışması için gerekli
        }
        return render(request, 'admin/import_courses_form.html', context)

@staff_member_required # Sadece admin/staff erişebilsin
def trigger_schedule_generation(request):
    if request.method == 'POST': # Sadece POST isteğiyle tetiklensin
        messages.info(request, "Ders programı oluşturma işlemi başlatıldı... Bu işlem uzun sürebilir.")
        
        try:
            scheduler = BacktrackingScheduler()
            success = scheduler.generate_and_save() # Algoritmayı çalıştır
            
            if success:
                messages.success(request, "Ders programı başarıyla oluşturuldu ve kaydedildi!")
            else:
                messages.error(request, "Ders programı oluşturulamadı veya tamamlanamadı. Detaylar için sunucu loglarına bakın.")
        except Exception as e:
            messages.error(request, f"Program oluşturma sırasında beklenmedik bir hata oluştu: {e}")
            
        # İşlem bittikten sonra admin ana sayfasına veya slot listesine yönlendir
        # Slot listesi daha mantıklı olabilir
        return redirect(reverse('admin:schedule_dersprogramislotu_changelist')) 
    else:
        # GET isteğiyle doğrudan gelinirse (veya POST değilse)
        # Belki bir onay sayfası gösterilebilir veya direkt ana sayfaya yönlendirilebilir
        return redirect(reverse('admin:index'))

@staff_member_required
def view_schedule(request):
    # GET parametrelerinden filtreleri al
    selected_bolum_id = request.GET.get('bolum')
    selected_sinif = request.GET.get('sinif')
    selected_semester = request.GET.get('semester') # Dönem filtresini al

    # Tüm bölümleri filtre dropdown için al
    all_bolumler = Bolum.objects.all()
    # Sınıfları tanımla (eğer sabitse)
    all_siniflar = SINIFLAR_DISPLAY # Dosyanın başındaki sabit listeyi kullanalım
    all_semesters = ["Güz", "Bahar"] # Dönem seçenekleri

    # Temel sorguyu oluştur
    slots = DersProgramiSlotu.objects.select_related('ders', 'ogretim_uyesi', 'derslik', 'bolum').order_by('gun', 'baslangic_saati', 'bolum', 'sinif')

    # Filtreleri uygula
    current_bolum = None
    bolum_adi_baslik = "Tüm Bölümler"
    if selected_bolum_id:
        try:
            slots = slots.filter(bolum_id=selected_bolum_id)
            current_bolum = all_bolumler.get(id=selected_bolum_id)
            bolum_adi_baslik = current_bolum.bolum_adi
        except (Bolum.DoesNotExist, ValueError):
             selected_bolum_id = None # Geçersiz ID ise filtreyi temizle

    if selected_sinif:
        try:
            sinif_int = int(selected_sinif)
            if sinif_int in all_siniflar:
                slots = slots.filter(sinif=sinif_int)
                if current_bolum:
                     bolum_adi_baslik += f" - {sinif_int}. Sınıf"
                else:
                     bolum_adi_baslik = f"{sinif_int}. Sınıf (Tüm Bölümler)"
            else:
                selected_sinif = None # Geçersiz sınıfsa filtreyi temizle
        except ValueError: # Sınıf sayı değilse
            selected_sinif = None
    elif current_bolum: # Sadece bölüm seçilmişse başlığa ekle
         bolum_adi_baslik += " (Tüm Sınıflar)"
         
    # Dönem filtresini uygula
    if selected_semester and selected_semester in all_semesters:
        slots = slots.filter(semester=selected_semester)
        # Başlığa dönem bilgisini ekle
        if bolum_adi_baslik == "Tüm Bölümler":
            bolum_adi_baslik = f"{selected_semester} Dönemi (Tüm Bölümler)"
        elif selected_sinif: # Eğer sınıf da seçiliyse
             bolum_adi_baslik += f" - {selected_semester}"
        elif current_bolum: # Sadece bölüm seçiliyse
             bolum_adi_baslik = f"{current_bolum.bolum_adi} - {selected_semester} Dönemi"
    else:
        selected_semester = None # Geçersiz veya boşsa filtreyi temizle


    # Veriyi grid formatına hazırla
    schedule_grid = {} # Standart dict kullanalım
    # Filtrelenmiş slotlar üzerinden gridi doldur
    for slot in slots:
        saat_str = f"{slot.baslangic_saati.strftime('%H:%M')}-{slot.bitis_saati.strftime('%H:%M')}"
        # Dönem bilgisini ekleyelim
        ders_info = f"{slot.ders.ders_kodu}<br>{slot.ogretim_uyesi.ad_soyad}<br>({slot.derslik.derslik_adi})<br><b>({slot.semester})</b>" 
        
        # İç içe sözlükleri manuel oluşturalım
        if slot.gun not in schedule_grid:
            schedule_grid[slot.gun] = {}
        if saat_str not in schedule_grid[slot.gun]:
            schedule_grid[slot.gun][saat_str] = {}

        # Sadece ders bilgisini değil, ID'yi de sakla
        schedule_grid[slot.gun][saat_str][slot.sinif] = {'info': ders_info, 'id': slot.id}

    # Zaman dilimlerini string formatına çevirelim
    time_slots_str = [f"{ts[0].strftime('%H:%M')}-{ts[1].strftime('%H:%M')}" for ts in TIME_SLOTS_DISPLAY]

    # Template'de gösterilecek sınıfları belirle (filtreye göre)
    siniflar_to_display = []
    if selected_sinif:
        try:
            sinif_int = int(selected_sinif)
            if sinif_int in all_siniflar:
                 siniflar_to_display = [sinif_int]
            else: # Geçersiz sınıf seçilmişse tümünü göster
                 siniflar_to_display = all_siniflar
        except ValueError: # Sayısal olmayan sınıf değeri gelirse tümünü göster
            siniflar_to_display = all_siniflar
    else: # Sınıf filtresi yoksa tümünü göster
        siniflar_to_display = all_siniflar

    context = {
        'title': 'Haftalık Ders Programı',
        'schedule_grid': schedule_grid,
        'days': DAYS_DISPLAY,
        'time_slots': time_slots_str,
        'siniflar': all_siniflar, # Filtre dropdown için tüm sınıflar
        'siniflar_to_display': siniflar_to_display, # Tabloda gösterilecek sınıflar
        'bolum_adi_baslik': bolum_adi_baslik,
        'all_bolumler': all_bolumler, # Filtre için bölümleri gönder
        'selected_bolum_id': selected_bolum_id,
        'selected_sinif': selected_sinif,
        'selected_semester': selected_semester, # Seçili dönemi gönder
        'all_semesters': all_semesters, # Dönem filtresi için seçenekleri gönder
        'has_permission': True,
    }
    return render(request, 'admin/schedule/view_schedule.html', context)

@require_POST # Sadece POST isteklerini kabul et
@staff_member_required # Sadece yetkili kullanıcılar
def update_slot_position(request):
    try:
        data = json.loads(request.body)
        slot_id = data.get('slot_id')
        new_day = data.get('new_day')
        new_time_str = data.get('new_time_str') # "HH:MM-HH:MM" formatında

        if not all([slot_id, new_day, new_time_str]):
            return JsonResponse({'success': False, 'error': 'Eksik veri.'}, status=400)

        # Zaman string'ini ayır ve time objelerine çevir
        try:
            start_time_str, end_time_str = new_time_str.split('-')
            new_start_time = datetime.datetime.strptime(start_time_str, '%H:%M').time()
            new_end_time = datetime.datetime.strptime(end_time_str, '%H:%M').time()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Geçersiz zaman formatı.'}, status=400)

        # İlgili slotu bul ve güncelle
        try:
            slot = DersProgramiSlotu.objects.get(id=slot_id)
            slot.gun = new_day
            slot.baslangic_saati = new_start_time
            slot.bitis_saati = new_end_time
            # TODO: Çakışma kontrolü burada eklenebilir (isteğe bağlı ama önerilir)
            # Örnek: Yeni konumda başka bir ders var mı?
            # existing_slot = DersProgramiSlotu.objects.filter(
            #     gun=new_day,
            #     baslangic_saati=new_start_time,
            #     # Çakışmayı daha detaylı kontrol etmek gerekebilir (derslik, hoca vs.)
            # ).exclude(id=slot_id).first()
            # if existing_slot:
            #     return JsonResponse({'success': False, 'error': f'Yeni konumda başka bir ders var: {existing_slot}'}, status=400)

            slot.save()
            return JsonResponse({'success': True})

        except DersProgramiSlotu.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Slot bulunamadı.'}, status=404)
        except Exception as e:
            # Beklenmedik veritabanı hataları vb.
            logger.error(f"Error updating slot {slot_id}: {e}")
            return JsonResponse({'success': False, 'error': f'Veritabanı hatası: {e}'}, status=500)

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Geçersiz JSON.'}, status=400)
    except Exception as e:
        logger.error(f"Unexpected error in update_slot_position: {e}")
        return JsonResponse({'success': False, 'error': 'Beklenmedik sunucu hatası.'}, status=500)
